from email import header
import pandas as pd
from flask import Flask
from flask_restful import Resource, Api, reqparse
from openpyxl import Workbook, load_workbook
import re
from flask import make_response

app = Flask(__name__)
api = Api(app)

data_arg = reqparse.RequestParser()
data_arg.add_argument('ID', type=int, required=True)
data_arg.add_argument('Novel', type=str, required=True)
data_arg.add_argument('Novellink', type=str, required=True)
data_arg.add_argument('Author', type=str, required=True)
data_arg.add_argument('AuthorLink', type=str, required=True)
data_arg.add_argument('Country', type=str, required=True)
data_arg.add_argument('CountryLink', type=str, required=True)


def make_hyperlink(url, value):
    return f'=HYPERLINK("{url}", "{value}")'


def retrieve_hyperlink_data(data_list):
    for dict in data_list:
        for col in ['الرواية', 'المؤلف', 'البلد']:
            match = re.search(r"=HYPERLINK\(\"(.*)\", \"(.*)\"\)", dict[col])
            if match.group(1):
                dict['رابط ' + col] = match.group(1)
            dict[col] = match.group(2)
    return data_list


class All(Resource):

    def get(self):
        global data
        data_list = data.to_dict('records')
        data_list = retrieve_hyperlink_data(data_list)
        return make_response({'data': data_list}, 200)


class read_Delete(Resource):

    # GET request on the url will hit this function
    def get(self, ID):
        global data
      # find data from xlsx based on user input
        data_list = data.loc[data['الترتيب'] == ID].to_dict('records')
        data_list = retrieve_hyperlink_data(data_list)
        # return data found in xlsx
        return make_response({'message': data_list}, 200)
    # Delete request on the url will hit this function

    def delete(self, ID):
        global data
        if ((data['الترتيب'] == ID).any()):
            # Id it present delete data from xlsx
            data = data.drop(data["الترتيب"].loc[data["الترتيب"] == ID].index)
            data.to_excel('Novels.xlsx', index=False)
            return make_response({"message": 'Deleted successfully'}, 202)
        else:
            return make_response({"message": 'Not Found'}, 204)


class Create_Update(Resource):

    # POST request on the url will hit this function
    def post(self):
        global data
        # data parser to parse data from url
        args = data_arg.parse_args()
        # if ID is already present
        if((data['الترتيب'] == args.ID).any()):
            return make_response({"message": 'ID already exist'}, 409)
        else:
            # Save data to xlsx
            data = data.append({
                "الترتيب": args.ID,
                "الرواية":  make_hyperlink(args.Novellink, args.Novel),
                "المؤلف": make_hyperlink(args.AuthorLink, args.Author),
                "البلد": make_hyperlink(args.CountryLink, args.Country),
            }, ignore_index=True)
            data = data.sort_values('الترتيب')
            data.to_excel('Novels.xlsx', index=False)
            return make_response({"message": 'Created successfully'}, 201)

    # PUT request on the url will hit this function
    def put(self):
        global data
        args = data_arg.parse_args()
        if ((data['الترتيب'] == args.ID).any()):
            # if ID already present Update it
            data = data.drop(
                data["الترتيب"].loc[data["الترتيب"] == args.ID].index)
            data = data.append({
                "الترتيب": args.ID,
                "الرواية":  make_hyperlink(args.Novellink, args.Novel),
                "المؤلف": make_hyperlink(args.AuthorLink, args.Author),
                "البلد": make_hyperlink(args.CountryLink, args.Country),
            }, ignore_index=True)
            data = data.sort_values('الترتيب')
            data.to_excel('Novels.xlsx', index=False)
            return make_response({"message": 'Updated successfully'}, 202)
        else:
            # If ID not present Save that data to xlsx
            data = data.append({
                "الترتيب": args.ID,
                "الرواية":  make_hyperlink(args.Novellink, args.Novel),
                "المؤلف": make_hyperlink(args.AuthorLink, args.Author),
                "البلد": make_hyperlink(args.CountryLink, args.Country),
            }, ignore_index=True)
            data = data.sort_values('الترتيب')
            data.to_excel('Novels.xlsx', index=False)
            return make_response({"message": 'Created successfully'}, 201)


# Add URL endpoints
api.add_resource(All, '/novels')
api.add_resource(read_Delete, '/novel/<int:ID>')
api.add_resource(Create_Update, '/novel')

if __name__ == '__main__':

    datawb = Workbook()
    datawb = load_workbook('Novels.xlsx')
    data = datawb.active
    data = pd.DataFrame(data.values, columns=[col.value for col in data[1]])
    data.drop(index=data.index[0], axis=0, inplace=True)
    data = data.astype({'الترتيب': 'int'})

    app.run(debug=True)